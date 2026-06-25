import os
import openpyxl
from pymongo import MongoClient
from dotenv import load_dotenv

load_dotenv()

def seed_db():
    excel_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "Question Bank - Behavior Round 1.xlsx"))
    print(f"Reading excel from: {excel_path}")
    
    if not os.path.exists(excel_path):
        print(f"Excel file not found at {excel_path}")
        return
        
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    print("Sheets found:", wb.sheetnames)
    
    if "FreeAssessment" not in wb.sheetnames:
        print("FreeAssessment sheet not found!")
        return
        
    sheet = wb["FreeAssessment"]
    
    # Read rows
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        print("No rows found in sheet")
        return
        
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    print("Headers:", headers)
    
    # Find indices of required columns
    col_map = {name: i for i, name in enumerate(headers) if name}
    
    # Connect to MongoDB
    mongo_uri = os.getenv("MONGO_URI", "mongodb://localhost:27017")
    print(f"Connecting to MongoDB at {mongo_uri}")
    client = MongoClient(mongo_uri)
    db = client["interview_db"]
    
    # Drop existing questions collection to reseed clean data
    db["questions"].drop()
    print("Dropped existing questions collection")
    
    records = []
    for r_idx, row in enumerate(rows[1:], start=2):
        # Skip empty rows
        if not any(cell is not None for cell in row):
            continue
            
        def get_val(col_name, default=""):
            idx = col_map.get(col_name)
            if idx is not None and idx < len(row):
                val = row[idx]
                return str(val).strip() if val is not None else default
            return default
            
        try:
            q_id_str = get_val("assessmentQuestionID")
            if not q_id_str:
                continue
            q_id = int(float(q_id_str)) # Handle cases like '73.0'
            
            q_num_str = get_val("assessmentQuestionNumber")
            q_num = int(float(q_num_str)) if q_num_str else 1
        except Exception as e:
            print(f"Row {r_idx}: Failed to parse questionID/number: {e}")
            continue
            
        records.append({
            "questionID": q_id,
            "question": get_val("question"),
            "companyApplicable": get_val("companyApplicable"),
            "roleFamily": get_val("roleFamily"),
            "roleLevel": get_val("roleLevel"),
            "roleTrack": get_val("roleTrack"),
            "prepMode": get_val("assessmentMode"),
            "assessmentQuestionNumber": q_num,
            "questionCompetency": get_val("questionCompetency")
        })
        
    if records:
        result = db["questions"].insert_many(records)
        print(f"Successfully inserted {len(result.inserted_ids)} assessment questions into MongoDB!")
    else:
        print("No records found to insert.")
        
    # Seed prompt
    prompt_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "scoring_prompt_text.txt"))
    print(f"Reading scoring prompt from: {prompt_path}")
    if os.path.exists(prompt_path):
        with open(prompt_path, "r", encoding="utf-8") as f:
            prompt_content = f.read()
        db["prompts"].drop()
        print("Dropped existing prompts collection")
        db["prompts"].insert_one({
            "key": "scoring_prompt",
            "content": prompt_content
        })
        print("Successfully inserted scoring prompt into MongoDB!")
    else:
        print(f"Scoring prompt file not found at {prompt_path}")
        
    client.close()

if __name__ == "__main__":
    seed_db()
